import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as graph_objects
from datetime import datetime, timedelta
from PIL import Image
import os
import io
import json
import unicodedata

# Nombre oficial del archivo base y carpeta física del proyecto
ARCHIVO_DB = "base_matriz_mce.xlsx"
CARPETA_EVIDENCIAS = "evidencias"
CATALOGO_FILE = "catalogos.json"

if not os.path.exists(CARPETA_EVIDENCIAS):
    os.makedirs(CARPETA_EVIDENCIAS)

Image.MAX_IMAGE_PIXELS = None

# Helper para normalizar texto (elimina acentos y convierte a minúsculas)
def normalizar_texto(texto):
    if not texto:
        return ""
    return "".join(
        c for c in unicodedata.normalize('NFD', str(texto).lower().strip())
        if unicodedata.category(c) != 'Mn'
    )

# Helper para cargar catálogos desde catalogos.json con fallbacks corregidos
def cargar_catalogos():
    if os.path.exists(CATALOGO_FILE):
        try:
            with open(CATALOGO_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                personal_list = data.get("personal", [])
                areas_list = data.get("areas", [])
                # Convertir a dict para compatibilidad con código original
                personal_dict = {name: None for name in personal_list}
                return personal_dict, areas_list
        except Exception as e:
            st.error(f"Error cargando catalogos.json: {e}")
            
    # Valores por defecto corregidos ortográficamente
    personal_def = {
        "Jesús Morales": None, "Cruz Carreón": None, "Luis Quintana": None, "Bryan Flores": None, "Rodolfo Fernández M.": None,
        "Ing. Alfredo Hernández": None, "Ing. Lorena Hernández": None, "Alejandra Arellano": None, "José Romo": None,
        "José Manuel Aldama": None, "Fernando Llanas": None, "Celso": None, "Josué Mesta": None, "Jorge Sánchez": None,
        "Víctor Montoya": None
    }
    areas_def = [
        "⚙️ Ingeniería", "🔍 Calidad", "📦 Almacén", "✂️ Corte", "📐 Doblez", "🎨 Pintura",
        "🚚 Embarques", "🏭 Planta Rio", "🛠️ Lijado", "💼 Administración", "👥 Recursos Humanos", "👑 Dirección"
    ]
    return personal_def, areas_def

def guardar_catalogos(personal_dict, areas_list):
    try:
        personal_list = list(personal_dict.keys())
        with open(CATALOGO_FILE, "w", encoding="utf-8") as f:
            json.dump({"personal": personal_list, "areas": areas_list}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        st.error(f"Error guardando catálogos: {e}")

# Inicialización de catálogos en sesión
if 'personal' not in st.session_state or 'areas' not in st.session_state:
    p_dict, a_list = cargar_catalogos()
    if 'personal' not in st.session_state:
        st.session_state.personal = p_dict
    if 'areas' not in st.session_state:
        st.session_state.areas = a_list

# Inicialización del estado de autenticación
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.rol = None
    st.session_state.usuario_actual = None

# 1. MOTOR DE IMPORTACIÓN INTELIGENTE DE EXCEL
def importar_registros_excel():
    if os.path.exists(ARCHIVO_DB):
        try:
            df = pd.read_excel(ARCHIVO_DB)
            if not df.empty:
                # CORRECCIÓN DE FECHAS: Detecta formatos datetime extensos y los limpia a texto DD-Mes-YY
                for col_fecha in ["Fecha Inicio", "Fecha Compromiso"]:
                    if col_fecha in df.columns:
                        def corregir_fecha_serial(val):
                            try:
                                if pd.isna(val) or str(val).strip() in ["None", "nan", "NaN", ""]: 
                                    return ""
                                if isinstance(val, (datetime, pd.Timestamp)):
                                    return val.strftime("%d-%b-%y")
                                if " " in str(val):
                                    val = str(val).split(" ")[0]
                                if str(val).replace('.0', '').isdigit():
                                    dias = int(str(val).replace('.0', ''))
                                    return (datetime(1899, 12, 30) + timedelta(days=dias)).strftime("%d-%b-%y")
                                return str(val).strip()
                            except: 
                                return str(val)
                        df[col_fecha] = df[col_fecha].apply(corregir_fecha_serial)

                if "% Avance" in df.columns:
                    if df["% Avance"].max() <= 1.0 and df["% Avance"].max() > 0: df["% Avance"] = df["% Avance"] * 100
                    df["% Avance"] = pd.to_numeric(df["% Avance"], errors="coerce").fillna(0).astype(int)
                
                if "No" in df.columns:
                    df["No"] = pd.to_numeric(df["No"], errors="coerce")
                    if df["No"].isnull().any(): df["No"] = range(1, len(df) + 1)
                    df["No"] = df["No"].astype(int)
                
                columnas_texto = ["Origen", "Prioridad", "Responsable", "Area", "Descripcion", "Comentario", "Evidencia"]
                for col in columnas_texto:
                    if col in df.columns: df[col] = df[col].astype(str).replace(["None", "nan", "NaN"], "")
            else:
                df = pd.DataFrame(columns=["No", "Origen", "Fecha Inicio", "Prioridad", "Responsable", "Area", "Descripcion", "% Avance", "Fecha Compromiso", "Comentario", "Evidencia"])
            return df
        except Exception as e:
            st.error(f"Error al importar el archivo Excel: {e}")
            return pd.DataFrame(columns=["No", "Origen", "Fecha Inicio", "Prioridad", "Responsable", "Area", "Descripcion", "% Avance", "Fecha Compromiso", "Comentario", "Evidencia"])
    else:
        return pd.DataFrame(columns=["No", "Origen", "Fecha Inicio", "Prioridad", "Responsable", "Area", "Descripcion", "% Avance", "Fecha Compromiso", "Comentario", "Evidencia"])

if 'actividades' not in st.session_state:
    st.session_state.actividades = importar_registros_excel()

# Limpiador de texto unicode para PDF (evitar caídas en FPDF2)
def limpiar_para_pdf(texto):
    if not texto:
        return ""
    texto = str(texto)
    reemplazos = {
        "\u201c": '"', "\u201d": '"', "\u2018": "'", "\u2019": "'",
        "\u2014": "-", "\u2013": "-", "\u2022": "*", "\u2026": "..."
    }
    for char, reemplazo in reemplazos.items():
        texto = texto.replace(char, reemplazo)
    return texto.encode('latin-1', 'replace').decode('latin-1')

st.set_page_config(page_title="SIGRAMA - Matriz MCE", layout="wide")

# INYECCIÓN CSS INTEGRAL DE IMAGEN CORPORATIVA (Industria SIGRAMA)
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;700&family=Questrial&display=swap');

    /* Fuentes globales */
    html, body, [class*="css"], .stApp {
        font-family: 'Questrial', sans-serif !important;
        background-color: #FFFFFF !important;
    }

    h1, h2, h3, h4, h5, h6, .main-title {
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 700 !important;
        color: #111111 !important;
    }

    /* Barra lateral corporativa en Negro profundo #111111 */
    [data-testid="stSidebar"] {
        background-color: #111111 !important;
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p, 
    [data-testid="stSidebar"] span, 
    [data-testid="stSidebar"] label {
        color: #FFFFFF !important;
        font-family: 'Questrial', sans-serif !important;
    }
    
    /* Logo negativo en sidebar */
    [data-testid="stSidebar"] img {
        filter: grayscale(1) invert(1) brightness(1.2) contrast(1.2) !important;
    }

    /* Botones de navegación en barra lateral */
    [data-testid="stSidebar"] div[role="radiogroup"] label {
        color: #FFFFFF !important;
        font-size: 14px !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] label:hover {
        color: #EC2024 !important;
    }

    /* Tarjetas personalizadas para actividades */
    .task-card {
        background-color: #FFFFFF;
        border: 1px solid #D2D3D5;
        border-radius: 6px;
        padding: 16px;
        margin-bottom: 12px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }

    /* Estilo de Botones Oficiales - Rojo Corporativo #EC2024 */
    div.stButton > button {
        background-color: #EC2024 !important;
        color: #FFFFFF !important;
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 700 !important;
        border-radius: 4px !important;
        border: 1px solid #EC2024 !important;
        padding: 8px 20px !important;
        transition: all 0.3s ease !important;
        text-transform: uppercase;
        font-size: 13px !important;
    }
    div.stButton > button:hover {
        background-color: #FFFFFF !important;
        color: #EC2024 !important;
        border: 1px solid #EC2024 !important;
    }

    /* Tarjetas de Métricas */
    [data-testid="metric-container"] {
        background-color: #FFFFFF !important;
        border: 1px solid #D2D3D5 !important;
        border-left: 5px solid #EC2024 !important;
        border-radius: 4px !important;
        padding: 12px 18px !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05) !important;
    }
    [data-testid="metric-container"] label {
        font-family: 'Montserrat', sans-serif !important;
        color: #111111 !important;
        font-weight: 500 !important;
    }
    [data-testid="metric-container"] div[data-testid="stMetricValue"] {
        font-family: 'Montserrat', sans-serif !important;
        color: #EC2024 !important;
        font-weight: 700 !important;
    }
    
    /* Configuración del Editor de Datos y Tablas */
    .stTable header, th {
        background-color: #111111 !important;
        color: #FFFFFF !important;
        font-family: 'Montserrat', sans-serif !important;
    }
    
    /* Inputs y Selectores */
    div[data-baseweb="input"], div[data-baseweb="select"], textarea {
        border-color: #D2D3D5 !important;
    }
    div[data-baseweb="input"]:focus-within, div[data-baseweb="select"]:focus-within {
        border-color: #EC2024 !important;
    }

    /* CONFIGURACIÓN MÁSTER EXCLUSIVA: SELECTOR GIGANTE DE OPERADOR PARA TOUCH SCREEN EN PISO */
    div[data-testid="stSidebar"] ~ div div[class*="stSelectbox"] div[data-baseweb="select"] {
        font-size: 32px !important; 
        font-weight: 800 !important; 
        color: #EC2024 !important;  
        height: 75px !important; 
        min-height: 75px !important;
        display: flex !important;
        align-items: center !important;
        border: 2px solid #EC2024 !important;
        border-radius: 6px !important;
    }
    
    div[data-testid="stSidebar"] ~ div div[class*="stSelectbox"] [data-testid="stSelectbox-SingleValue"],
    div[data-testid="stSidebar"] ~ div div[class*="stSelectbox"] div[data-baseweb="select"] span {
        line-height: 75px !important;
        font-size: 32px !important;
        overflow: visible !important; 
        font-family: 'Montserrat', sans-serif !important;
        color: #EC2024 !important;
    }

    /* Calibración de la lista desplegable de nombres gigante */
    div[data-baseweb="popover"] ul li {
        font-size: 24px !important;
        padding-top: 8px !important;
        padding-bottom: 8px !important;
        line-height: 1.2 !important;
    }

    /* Mantener selectores normales en tablas de datos */
    div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"],
    div[data-testid="element-container"] div[data-testid="stSelectbox"] div[data-baseweb="select"] {
        font-size: 14px !important;
        height: auto !important;
        min-height: auto !important;
        border: 1px solid #D2D3D5 !important;
    }
    </style>
""", unsafe_allow_html=True)

LISTA_CLASIFICACIONES = ["Acuerdos", "Programa de Actividades", "Actividades Sujeridas", "Dirección", "Problema de Calidad", "Problema de Seguridad", "Lista de Pendientes", "Auto Asignado", "Plan de Control y Monitoreo", "Mejoras", "Investigación", "Manuales", "Procesos"]

def crear_grafico_pareto(df, columna, titulo):
    if df.empty:
        fig = graph_objects.Figure(); fig.update_layout(title=f"{titulo} (Sin Datos)"); return fig
    df = df.copy()
    df["Estado"] = df["% Avance"].apply(lambda x: "Terminada" if x == 100 else "Pendiente")
    orden_cat = df[columna].value_counts().index.tolist()
    counts = df[columna].value_counts().reset_index()
    counts.columns = [columna, 'Cantidad']
    counts['Porcentaje Acumulado'] = (counts['Cantidad'].cumsum() / len(df)) * 100
    
    # Rojo oficial #EC2024 para pendientes y verde #2ECC71 para terminadas
    fig = px.histogram(df, x=columna, color="Estado", category_orders={columna: orden_cat}, color_discrete_map={"Terminada": "#2ECC71", "Pendiente": "#EC2024"}, title=titulo)
    fig.add_trace(graph_objects.Scatter(x=counts[columna], y=counts['Porcentaje Acumulado'], name="% Acumulado", yaxis="y2", line=dict(color="#FF8A00", width=3)))
    fig.update_layout(yaxis=dict(title="Cantidad"), yaxis2=dict(title="% Acumulado", overlaying="y", side="right", range=[0, 100]), legend=dict(orientation="h", y=1.02, x=1, xanchor="right"), template="plotly_white", barmode="stack")
    return fig

# 5. Carga e Inyección del Banner Corporativo Superior
nombre_banner = "LIDERAZGO PLANTA METALAES IMAGEN.png"
if os.path.exists(nombre_banner):
    imagen_banner = Image.open(nombre_banner)
    st.image(imagen_banner, use_container_width=True)
else:
    st.markdown('<h2 style="color: #EC2024; text-align: center; font-weight: bold; margin-top:0px; font-family: \'Montserrat\', sans-serif;">PLANTA METALES Y MAQUINADOS</h2>', unsafe_allow_html=True)
    st.markdown('<p style="text-align: center; font-size: 24px; font-weight: bold; color: #111111; font-family: \'Montserrat\', sans-serif;">MATRIZ DE COMUNICACIÓN EFECTIVA</p>', unsafe_allow_html=True)

# ----------------- FLUJO DE AUTENTICACIÓN (LOGIN) -----------------
if not st.session_state.logged_in:
    # Sidebar sin navegación cuando no ha iniciado sesión
    if os.path.exists("LOGOTIPO COLOR (1).jfif"):
        st.sidebar.image("LOGOTIPO COLOR (1).jfif", use_container_width=True)
        
    st.sidebar.markdown("""
    <div style="background-color: #0F172A; border: 1px solid #1E293B; padding: 15px; border-radius: 6px; text-align: center; margin-top: 20px;">
        <p style="color: #F8FAFC; font-family: 'Questrial', sans-serif; font-size: 13px; font-weight: 500; margin: 0; line-height: 1.4;">
            <span style="color: #F59E0B;">🔒</span> Por favor, inicie sesión en la pantalla principal para configurar la aplicación.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    st.sidebar.markdown("""
        <div style="text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #D2D3D5;">
            <span style="font-family: 'Questrial', sans-serif; font-style: italic; font-size: 13px; color: #FFFFFF; border-bottom: 2px solid #EC2024; padding-bottom: 4px; display: inline-block;">
                Ingeniería que da resultados!!
            </span>
        </div>
    """, unsafe_allow_html=True)

    # Vista Principal de Inicio de Sesión
    st.markdown('<p style="text-align: center; font-size: 16px; font-weight: bold; color: #EC2024; font-family: \'Montserrat\', sans-serif; margin-top: 15px; text-transform: uppercase; letter-spacing: 1px;">SOLUCIONES QUE TRANSFORMAN TU EMPRESA</p>', unsafe_allow_html=True)
    st.markdown('<hr style="border: 1px solid #EC2024; margin: 15px 0;">', unsafe_allow_html=True)
    st.markdown("""
    <h2 style="text-align: center; font-family: 'Montserrat', sans-serif; color: #111111; font-weight: 700; font-size: 28px; margin: 10px 0;">
        <span style="color: #EC2024;">📢</span> Matriz de Comunicación Efectiva
    </h2>
    """, unsafe_allow_html=True)
    st.markdown('<hr style="border: 1px solid #EC2024; margin: 15px 0;">', unsafe_allow_html=True)
    
    col_log1, col_log2, col_log3 = st.columns([1, 2, 1])
    with col_log2:
        st.markdown('<div style="background-color: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 8px; padding: 25px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -1px rgba(0,0,0,0.06); margin-top: 20px;">', unsafe_allow_html=True)
        st.markdown('<h3 style="font-family: \'Montserrat\', sans-serif; font-weight: 700; color: #111111; text-align: center; margin-top: 0; font-size: 20px;"><span style="color: #F59E0B;">🔑</span> Acceso al Sistema de Control</h3>', unsafe_allow_html=True)
        st.markdown('<p style="font-family: \'Questrial\', sans-serif; color: #64748B; font-size: 14px; text-align: center; margin-bottom: 20px;">Por favor, ingrese sus credenciales para operar la Matriz de Comunicación Efectiva.</p>', unsafe_allow_html=True)
        
        username_input = st.text_input("Usuario:", key="login_username")
        password_input = st.text_input("Contraseña:", type="password", key="login_password")
        
        st.markdown('<div style="margin-top: 20px;">', unsafe_allow_html=True)
        btn_login = st.button("Ingresar", use_container_width=True, key="btn_login_submit")
        st.markdown('</div></div>', unsafe_allow_html=True)
        
        if btn_login:
            username_norm = normalizar_texto(username_input)
            colaborador_encontrado = None
            for name in st.session_state.personal.keys():
                if normalizar_texto(name) == username_norm:
                    colaborador_encontrado = name
                    break
            
            if username_norm in ["admin", "administrador"] and password_input == "SigramaMetales2026":
                st.session_state.logged_in = True
                st.session_state.rol = "Administrador"
                st.session_state.usuario_actual = "Administrador"
                st.success("Sesión iniciada como Administrador.")
                st.rerun()
            elif colaborador_encontrado is not None and password_input == "Metales":
                st.session_state.logged_in = True
                st.session_state.rol = "Colaborador"
                st.session_state.usuario_actual = colaborador_encontrado
                st.success(f"Sesión iniciada como {colaborador_encontrado}.")
                st.rerun()
            else:
                st.error("Credenciales incorrectas. Verifique el usuario y la contraseña.")

else:
    # ----------------- MENU PRINCIPAL Y NAVEGACIÓN (LOGGED IN) -----------------
    if os.path.exists("LOGOTIPO COLOR (1).jfif"):
        st.sidebar.image("LOGOTIPO COLOR (1).jfif", use_container_width=True)

    st.sidebar.markdown(f"""
    <div style="background-color: #0F172A; border: 1px solid #1E293B; padding: 10px; border-radius: 6px; text-align: center; margin-bottom: 15px;">
        <span style="color: #F8FAFC; font-family: 'Questrial', sans-serif; font-size: 13px; font-weight: 500;">
            👤 Usuario: <b>{st.session_state.usuario_actual}</b>
        </span>
    </div>
    """, unsafe_allow_html=True)

    # Filtrar navegación de acuerdo al rol de seguridad
    menu_options = [
        "📊 Dashboard Principal", 
        "📋 Tabla de Control", 
        "📝 Actualizar Mis Avances", 
        "📥 Cargar Actividades (Usuario)", 
        "👑 Reglas de Liderazgo", 
        "📋 Reportes PDF",
        "🏭 Industria 4.0 & Stack"
    ]
    if st.session_state.rol == "Administrador":
        menu_options.insert(4, "🔐 Panel Administrador")

    opcion_menu = st.sidebar.radio("Navegación", menu_options)

    # Botón de Cerrar Sesión en Sidebar
    if st.sidebar.button("🚪 Cerrar Sesión", use_container_width=True):
        st.session_state.logged_in = False
        st.session_state.rol = None
        st.session_state.usuario_actual = None
        st.rerun()

    # Firma al final de la navegación
    st.sidebar.markdown("""
        <div style="text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #D2D3D5;">
            <span style="font-family: 'Questrial', sans-serif; font-style: italic; font-size: 13px; color: #FFFFFF; border-bottom: 2px solid #EC2024; padding-bottom: 4px; display: inline-block;">
                Ingeniería que da resultados!!
            </span>
        </div>
    """, unsafe_allow_html=True)

    # ----------------- RENDERIZACIÓN DE SECCIONES -----------------
    if opcion_menu == "📊 Dashboard Principal":
        col_f1, col_f2 = st.columns(2)
        with col_f1: area_sel = st.selectbox("Filtrar por Área", ["Todas"] + st.session_state.areas)
        with col_f2: resp_sel = st.selectbox("Filtrar por Responsable", ["Todos"] + list(st.session_state.personal.keys()))
        df_f = pd.DataFrame(st.session_state.actividades)
        if area_sel != "Todas": df_f = df_f[df_f["Area"] == area_sel]
        if resp_sel != "Todos": df_f = df_f[df_f["Responsable"] == resp_sel]
        
        c1, c2, c3 = st.columns(3)
        c1.metric("Total Actividades", len(df_f))
        c2.metric("Terminadas", len(df_f[df_f["% Avance"] == 100]))
        c3.metric("Avance Promedio", f"{df_f['% Avance'].mean() if len(df_f)>0 else 0:.1f}%")
        st.write("---")
        g1, g2 = st.columns(2)
        with g1: st.plotly_chart(crear_grafico_pareto(df_f, "Origen", "Pareto 1: Actividades vs Cantidad"), use_container_width=True)
        with g2: st.plotly_chart(crear_grafico_pareto(df_f, "Responsable", "Pareto 2: Personas vs Cantidad"), use_container_width=True)
        
        st.write("---"); st.subheader("Pareto 3: Estado de Actividades de Líderes Principales")
        lideres_p = ["Jesús Morales", "Bryan Flores", "Cruz Carreón", "Luis Quintana"]
        df_lideres = pd.DataFrame(st.session_state.actividades)[lambda x: x["Responsable"].isin(lideres_p)].copy()
        if not df_lideres.empty:
            hoy = datetime.now()
            def clasificar_vencimientos(f):
                try:
                    f_comp = datetime.strptime(str(f["Fecha Compromiso"]).strip(), "%d-%b-%y")
                    return "Vencida (Retraso)" if (int(f["% Avance"]) < 100 and f_comp < hoy) else ("Terminada" if int(f["% Avance"]) == 100 else "Pendiente a Tiempo")
                except: return "Pendiente a Tiempo"
            df_lideres["Estado_Real"] = df_lideres.apply(clasificar_vencimientos, axis=1)
            df_lideres["Valor_Eje"] = df_lideres["Estado_Real"].apply(lambda x: -1 if x == "Vencida (Retraso)" else 1)
            
            fig_l = px.bar(
                df_lideres, 
                x="Origen", 
                y="Valor_Eje", 
                color="Estado_Real", 
                facet_col="Responsable", 
                facet_col_wrap=2, 
                color_discrete_map={"Terminada": "#2ECC71", "Pendiente a Tiempo": "#FFE600", "Vencida (Retraso)": "#EC2024"}, 
                labels={
                    "Origen": "Clasificación", 
                    "Valor_Eje": "Tareas",
                    "No": "Actividad Número",
                    "Descripcion": "Descripción",
                    "Fecha Inicio": "Fecha Inicio",
                    "Fecha Compromiso": "Fecha Compromiso",
                    "% Avance": "Avance",
                    "Estado_Real": "Estado"
                },
                hover_data={
                    "Valor_Eje": False,
                    "No": True,
                    "Descripcion": True,
                    "Fecha Inicio": True,
                    "Fecha Compromiso": True,
                    "% Avance": True,
                    "Estado_Real": True
                }
            )
            fig_l.update_layout(barmode="stack", template="plotly_white", height=600, legend=dict(orientation="h", y=-0.2, x=0.5, xanchor="center"))
            fig_l.for_each_annotation(lambda a: a.update(text=f"<b>👤 {a.text.split('=')[-1].upper()}</b>", font=dict(size=14, color="#111111")))
            st.plotly_chart(fig_l, use_container_width=True)

    elif opcion_menu == "📋 Tabla de Control":
        st.subheader("Historial Completo de la Matriz de Comunicación")
        df_t = pd.DataFrame(st.session_state.actividades)
        st.write("---")
        ct1, ct2, ct3 = st.columns(3)
        with ct1: fa = st.selectbox("Filtrar por Área", ["Todas"] + st.session_state.areas, key="t_a")
        with ct2: fr = st.selectbox("Filtrar por Responsable", ["Todos"] + list(st.session_state.personal.keys()), key="t_r")
        with ct3: fp = st.selectbox("Filtrar por Prioridad", ["Todas", "Urgente", "Media", "Baja"])
        if fa != "Todas": df_t = df_t[df_t["Area"] == fa]
        if fr != "Todos": df_t = df_t[df_t["Responsable"] == fr]
        if fp != "Todas": df_t = df_t[df_t["Prioridad"] == fp]
        busq = st.text_input("🔍 Buscar por palabra clave en la descripción:")
        if busq: df_t = df_t[df_t["Descripcion"].str.contains(busq, case=False, na=False)]
        if not df_t.empty:
            df_mostrar = df_t[["No", "Origen", "Fecha Inicio", "Prioridad", "Responsable", "Area", "Descripcion", "% Avance", "Fecha Compromiso", "Comentario", "Evidencia"]].copy()
            hoy = datetime.now()
            
            def aplicar_colores_renglon(fila):
                try:
                    num_avance = int(fila["% Avance"])
                    fecha_comp_str = str(fila["Fecha Compromiso"]).strip()
                    fecha_vencida = False
                    if fecha_comp_str:
                        try:
                            f_comp = datetime.strptime(fecha_comp_str, "%d-%b-%y")
                            if f_comp < hoy: fecha_vencida = True
                        except: pass
                    if num_avance < 100 and fecha_vencida: 
                        return ['background-color: #F8D7DA; color: #721C24; font-weight: 500;'] * len(fila)
                    elif num_avance == 100: 
                        return ['background-color: #D4EDDA; color: #155724;'] * len(fila)
                    elif num_avance > 0: 
                        return ['background-color: #FFF3CD; color: #856404;'] * len(fila)
                except: pass
                return [''] * len(fila)

            st.dataframe(df_mostrar.style.apply(aplicar_colores_renglon, axis=1).format({"% Avance": "{:.0f}%"}), use_container_width=True, hide_index=True)
            
            def generar_excel_con_colores(df_local):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_local.to_excel(writer, index=False, sheet_name='Historial_MCE')
                    workbook = writer.book
                    worksheet = writer.sheets['Historial_MCE']
                    
                    for col in worksheet.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = col[0].column_letter
                        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                        
                    from openpyxl.styles import PatternFill, Font
                    fill_verde = PatternFill(start_color="D4EDDA", fill_type="solid")
                    fill_amarillo = PatternFill(start_color="FFF3CD", fill_type="solid")
                    fill_rojo = PatternFill(start_color="F8D7DA", fill_type="solid")
                    font_rojo = Font(color="721C24", bold=True)
                    
                    headers = [worksheet.cell(row=1, column=c).value for c in range(1, worksheet.max_column + 1)]
                    idx_avance = headers.index("% Avance") + 1 if "% Avance" in headers else 8
                    idx_compromiso = headers.index("Fecha Compromiso") + 1 if "Fecha Compromiso" in headers else 9
                    
                    for row_idx in range(2, worksheet.max_row + 1):
                        try:
                            avance_val = int(str(worksheet.cell(row=row_idx, column=idx_avance).value).replace('%','').strip())
                            fecha_comp_str = str(worksheet.cell(row=row_idx, column=idx_compromiso).value).strip()
                            
                            es_vencido = False
                            if avance_val < 100 and fecha_comp_str and fecha_comp_str != "None":
                                from datetime import datetime
                                if datetime.strptime(fecha_comp_str, "%d-%b-%y") < datetime.now():
                                    es_vencido = True
                            
                            for col_idx in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                if avance_val < 100 and es_vencido:
                                    cell.fill = fill_rojo; cell.font = font_rojo
                                elif avance_val == 100:
                                    cell.fill = fill_verde
                                elif avance_val > 0:
                                    cell.fill = fill_amarillo
                        except:
                            pass
                return output.getvalue()

            excel_data = generar_excel_con_colores(df_mostrar)
            st.download_button(
                label="📥 Descargar Reporte en Excel (.xlsx con Colores)",
                data=excel_data,
                file_name=f"Reporte_Matriz_MCE_{datetime.now().strftime('%d-%b-%y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    elif opcion_menu == "📝 Actualizar Mis Avances":
        st.subheader("Actualización de Avances de Tareas")
        
        # Filtro de responsable autodetectado y bloqueado para Colaboradores
        if st.session_state.rol == "Colaborador":
            u = st.session_state.usuario_actual
            st.markdown(f"<div style='background-color:#E2E8F0; padding:10px; border-radius:4px; margin-bottom:15px; font-weight:bold;'>👤 Colaborador Activo: {u}</div>", unsafe_allow_html=True)
        else:
            u = st.selectbox("Seleccionar Responsable:", list(st.session_state.personal.keys()))
        
        df_usuario = pd.DataFrame(st.session_state.actividades)
        if not df_usuario.empty and "Responsable" in df_usuario.columns:
            df_usuario = df_usuario[df_usuario["Responsable"] == u]
        
        if df_usuario.empty: 
            st.info(f"👤 {u} no tiene actividades pendientes asignadas en este momento.")
        else:
            st.markdown('<p style="font-size:16px; font-weight:bold; color:#111111; margin-bottom:5px; font-family:\'Montserrat\', sans-serif;">📊 Mi Rendimiento Actual</p>', unsafe_allow_html=True)
            col_dash1, col_dash2, col_dash3 = st.columns(3)
            
            tareas_pendientes = len(df_usuario[df_usuario["% Avance"].astype(int) < 100])
            tareas_hechas = len(df_usuario[df_usuario["% Avance"].astype(int) == 100])
            promedio_avance = df_usuario["% Avance"].astype(int).mean()
            
            col_dash1.metric(label="⏳ En Proceso", value=f"{tareas_pendientes} tareas")
            col_dash2.metric(label="✅ Terminadas", value=f"{tareas_hechas} tareas")
            col_dash3.metric(label="📈 Eficiencia Total", value=f"{promedio_avance:.1f}%")
            st.write("---")
            
            clasificacion = st.radio(
                "Filtrar mi lista por estatus:",
                ["En proceso (0% a 99%)", "Terminadas (100%)", "Ver Todas las Asignadas"],
                horizontal=True
            )
            
            if clasificacion == "En proceso (0% a 99%)":
                df_filtrado = df_usuario[df_usuario["% Avance"].astype(int) < 100]
            elif clasificacion == "Terminadas (100%)":
                df_filtrado = df_usuario[df_usuario["% Avance"].astype(int) == 100]
            else:
                df_filtrado = df_usuario.copy()
                
            st.write("---")
            
            if df_filtrado.empty:
                st.info(f"📋 No hay tareas en la clasificación '{clasificacion}' para {u}.")
            else:
                for idx in df_filtrado.index:
                    r = st.session_state.actividades.loc[idx]
                    
                    with st.container():
                        st.markdown(f"""
                        <div class="task-card">
                            <p style="font-family:'Montserrat', sans-serif; font-size:18px; font-weight:bold; color:#111111; margin:0 0 6px 0;">
                                📋 Actividad No. {r["No"]} | {r["Area"]} | Prioridad: <span style="color:#EC2024; font-weight:700;">{r["Prioridad"]}</span>
                            </p>
                            <p style="font-family:'Questrial', sans-serif; font-size:15px; color:#333333; margin:0 0 10px 0;">
                                <b>Descripción:</b> {r["Descripcion"]}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        col_izq, col_der = st.columns(2)
                        
                        with col_izq:
                            progreso_actual = int(r['% Avance'])
                            
                            color_progreso = "#2ECC71" if progreso_actual == 100 else "#EC2024"
                            st.markdown(f"""
                            <div style="font-family:'Montserrat', sans-serif; font-size:13px; font-weight:700; color:#111111; margin-bottom:4px;">
                                PROGRESO: {progreso_actual}%
                            </div>
                            <div style="width:100%; background-color:#D2D3D5; border-radius:4px; height:18px; overflow:hidden; margin-bottom:12px; border:1px solid #D2D3D5;">
                                <div style="width:{progreso_actual}%; background-color:{color_progreso}; height:100%; transition:width 0.4s ease-in-out;"></div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            nv_av = st.slider("Ajustar %:", min_value=0, max_value=100, value=progreso_actual, step=5, key=f"num_{r['No']}")
                        
                        with col_der:
                            comentario_limpio = "" if str(r['Comentario']).strip().lower() in ["nan", "none", ""] else str(r['Comentario'])
                            nv_co = st.text_input("Comentarios de bitácora:", value=comentario_limpio, key=f"c_{r['No']}")
                            
                            evidencia_guardada = str(r['Evidencia']).strip()
                            if evidencia_guardada and os.path.exists(evidencia_guardada):
                                st.image(Image.open(evidencia_guardada), width=130, caption="📸 Evidencia Actual")
                            
                            foto = st.file_uploader("Evidencia Fotográfica (Cierre 100%):", type=["jpg","png","jpeg","jfif"], key=f"i_{r['No']}") if nv_av == 100 else None
                            if foto: 
                                st.image(Image.open(foto), width=100, caption="Vista Previa")
                            
                            st.markdown('<div style="margin-top: 10px;">', unsafe_allow_html=True)
                            if st.button("Guardar Tarea", key=f"b_{r['No']}", use_container_width=True):
                                if nv_av == 100 and not foto and not evidencia_guardada: 
                                    st.error("⚠️ Faltan las fotografías físicas obligatorias para autorizar el cierre al 100%.")
                                else:
                                    ruta_foto_final = evidencia_guardada
                                    if foto is not None:
                                        try:
                                            img_abierta = Image.open(foto)
                                            if img_abierta.mode in ("RGBA", "P"): 
                                                img_abierta = img_abierta.convert("RGB")
                                            img_abierta.thumbnail((800, 600), Image.Resampling.LANCZOS)
                                            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                            nombre_archivo_final = f"MCE-{int(r['No']):03d}_evidencia_{stamp}.jpg"
                                            ruta_foto_final = os.path.join(CARPETA_EVIDENCIAS, nombre_archivo_final)
                                            img_abierta.save(ruta_foto_final, "JPEG", quality=65)
                                        except Exception as err_img: 
                                            st.error(f"Fallo al procesar captura: {err_img}")
                                    
                                    st.session_state.actividades.loc[idx, "% Avance"] = int(nv_av)
                                    st.session_state.actividades.loc[idx, "Comentario"] = str(nv_co)
                                    st.session_state.actividades.loc[idx, "Evidencia"] = str(ruta_foto_final)
                                    
                                    try:
                                        st.session_state.actividades.to_excel(ARCHIVO_DB, index=False)
                                        st.success("🏁 ¡Cambios registrados con éxito!"); st.rerun()
                                    except Exception as e_save: 
                                        st.error(f"Fallo de escritura en base física Excel: {e_save}")
                            st.markdown('</div>', unsafe_allow_html=True)
                        st.write("---")

    elif opcion_menu == "📥 Cargar Actividades (Usuario)":
        st.subheader("Captura de Nuevas Actividades")

        with st.form("form_usuario_carga"):
            o, p, r, a = st.selectbox("Origen", LISTA_CLASIFICACIONES), st.selectbox("Prioridad", ["Baja", "Media", "Urgente"]), st.selectbox("Responsable", list(st.session_state.personal.keys())), st.selectbox("Área", st.session_state.areas)
            d, f = st.text_area("Descripción"), st.date_input("Fecha Compromiso")
            if st.form_submit_button("Registrar Actividad"):
                # Calcular tiempo para ejecutar
                try:
                    hoy = datetime.now().date()
                    dias = (f - hoy).days
                    if dias > 0:
                        tiempo_ejecutar = f"{dias} días naturales"
                    elif dias == 0:
                        tiempo_ejecutar = "Hoy mismo"
                    else:
                        tiempo_ejecutar = f"Atrasada por {abs(dias)} días"
                except Exception as e:
                    tiempo_ejecutar = "No definido"

                n_id = int(st.session_state.actividades["No"].max() + 1) if not st.session_state.actividades.empty else 1
                n_f = {"No": n_id, "Origen": o, "Fecha Inicio": datetime.now().strftime("%d-%b-%y"), "Prioridad": p, "Responsable": r, "Area": a, "Descripcion": d, "% Avance": 0, "Fecha Compromiso": f.strftime("%d-%b-%y"), "Comentario": "", "Evidencia": ""}
                st.session_state.actividades = pd.concat([st.session_state.actividades, pd.DataFrame([n_f])], ignore_index=True)
                
                try:
                    st.session_state.actividades.to_excel(ARCHIVO_DB, index=False)
                    # Guardar resumen en session_state para mostrar tras rerun
                    st.session_state.last_registered_activity = {
                        "No": n_id,
                        "Fecha": datetime.now().strftime("%d-%b-%y"),
                        "Responsable": r,
                        "Nombre": d,
                        "Tiempo": tiempo_ejecutar
                    }
                    st.rerun()
                except Exception as e_add: st.error(f"Fallo Excel: {e_add}")

        # Mostrar resumen de validación si existe, debajo del formulario/botón
        if 'last_registered_activity' in st.session_state and st.session_state.last_registered_activity is not None:
            act = st.session_state.last_registered_activity
            st.markdown(f"""
            <div style="background-color: #D4EDDA; border: 1px solid #C3E6CB; color: #155724; padding: 18px; border-radius: 6px; margin-top: 15px; margin-bottom: 20px; font-family: 'Questrial', sans-serif;">
                <h4 style="margin-top: 0; color: #155724; font-family: 'Montserrat', sans-serif; font-weight: 700; font-size: 16px;">✅ ¡Actividad Registrada con Éxito y Validada!</h4>
                <hr style="border: 0.5px solid #C3E6CB; margin: 8px 0;">
                <p style="margin: 3px 0; font-size: 14px;"><b>No. Actividad:</b> {act['No']}</p>
                <p style="margin: 3px 0; font-size: 14px;"><b>Fecha de Registro:</b> {act['Fecha']}</p>
                <p style="margin: 3px 0; font-size: 14px;"><b>Responsable:</b> {act['Responsable']}</p>
                <p style="margin: 3px 0; font-size: 14px;"><b>Actividad (Descripción):</b> {act['Nombre']}</p>
                <p style="margin: 3px 0; font-size: 14px; color: #856404; font-weight: bold;">⏳ Tiempo Restante para Ejecutar: {act['Tiempo']}</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Limpiar Notificación", key="btn_clear_notif"):
                st.session_state.last_registered_activity = None
                st.rerun()
        
        st.write("---")
        st.subheader("📋 Lista de Actividades (Últimas Cargadas en Amarillo)")
        
        df_mostrar_carga = pd.DataFrame(st.session_state.actividades)
        if not df_mostrar_carga.empty:
            # Ordenar por No descendente (últimas cargadas primero)
            df_mostrar_carga = df_mostrar_carga.sort_values(by="No", ascending=False)
            
            # Highlight threshold: las últimas 5 ingresadas (No mayores o iguales al max(No) - 4)
            limite_no = df_mostrar_carga["No"].max() - 4
            
            def destacar_recientes(row):
                try:
                    if int(row["No"]) >= limite_no:
                        return ['background-color: #FFF3CD; color: #856404; font-weight: 500;'] * len(row)
                except:
                    pass
                return [''] * len(row)
                
            st.dataframe(df_mostrar_carga[["No", "Origen", "Fecha Inicio", "Prioridad", "Responsable", "Area", "Descripcion", "% Avance", "Fecha Compromiso"]].style.apply(destacar_recientes, axis=1), use_container_width=True, hide_index=True)
        else:
            st.info("No hay actividades registradas en el sistema.")

    elif opcion_menu == "🔐 Panel Administrador" and st.session_state.rol == "Administrador":
        st.markdown('### Panel de Control Máster')
        st.success("Acceso Máster Autorizado")
        st.write("---")
        st.markdown("### 📊 Consola de Sincronización Automática con GitHub (API REST)")
        c_adm1, c_adm2 = st.columns(2)
        with c_adm1:
            if st.button("📥 IMPORTAR BASE DE DATOS DESDE GITHUB", use_container_width=True, key="btn_import_master"):
                st.session_state.actividades = importar_registros_excel()
                st.success("¡Sincronizado! Datos actualizados desde GitHub."); st.rerun()
        with c_adm2:
            if st.button("🚀 RESPALDAR AND SUBIR DIRECTO A GITHUB", type="primary", use_container_width=True, key="btn_git_api_push_master"):
                import requests, base64, json
                try:
                    df_guardar = pd.DataFrame(st.session_state.actividades)
                    with pd.ExcelWriter(ARCHIVO_DB, engine='openpyxl') as w:
                        df_guardar.to_excel(w, index=False, sheet_name='Base_MCE')
                        ws = w.sheets['Base_MCE']
                        anchos = {'A': 10, 'B': 25, 'C': 15, 'D': 15, 'E': 22, 'F': 18, 'G': 45, 'H': 12, 'I': 20, 'J': 25, 'K': 40}
                        for col, ancho in anchos.items(): ws.column_dimensions[col].width = ancho
                        
                        from openpyxl.styles import PatternFill, Font
                        fill_verde, fill_amarillo, fill_rojo = PatternFill(start_color="D4EDDA", fill_type="solid"), PatternFill(start_color="FFF3CD", fill_type="solid"), PatternFill(start_color="F8D7DA", fill_type="solid")
                        font_rojo, font_normal = Font(color="721C24", bold=True), Font(color="000000")
                        hoy_dt = datetime.now()
                        
                        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                        idx_avance = headers.index("% Avance") + 1 if "% Avance" in headers else 8
                        idx_compromiso = headers.index("Fecha Compromiso") + 1 if "Fecha Compromiso" in headers else 9
                        
                        for row_idx in range(2, ws.max_row + 1):
                            try:
                                avance_val = int(str(ws.cell(row=row_idx, column=idx_avance).value).replace('%','').strip())
                                fecha_comp_str = str(ws.cell(row=row_idx, column=idx_compromiso).value).strip()
                                es_vencido = False
                                if avance_val < 100 and fecha_comp_str:
                                    if datetime.strptime(fecha_comp_str, "%d-%b-%y") < hoy_dt: es_vencido = True
                                for col_idx in range(1, ws.max_column + 1):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if avance_val < 100 and es_vencido: cell.fill = fill_rojo; cell.font = font_rojo
                                    elif avance_val == 100: cell.fill = fill_verde; cell.font = font_normal
                                    elif avance_val > 0: cell.fill = fill_amarillo; cell.font = font_normal
                            except: pass
                        
                        try:
                            token_git = st.secrets["TOKEN_GITHUB"]
                        except Exception:
                            token_git = ""
                
                        usuario_git = "jesusalbertomoraleslopez-byte"
                        repo_git = "matriz-mce-sigrama"
                        email_git = "jesusalbertomoraleslopez@gmail.com"
                
                        url_api = f"https://api.github.com/repos/{usuario_git}/{repo_git}/contents/base_matriz_mce.xlsx"
                        cabeceras = {"Authorization": f"token {token_git}", "Accept": "application/vnd.github.v3+json"}
                        respuesta_get = requests.get(url_api, headers=cabeceras)
                        sha_archivo = None
                        if respuesta_get.status_code == 200:
                            try:
                                sha_archivo = respuesta_get.json().get("sha")
                            except:
                                pass
                        elif respuesta_get.status_code != 404:
                            st.error(f"Fallo de comunicación con GitHub API (Código: {respuesta_get.status_code}).")
                            st.stop()

                    with open(ARCHIVO_DB, "rb") as archivo_binario: excel_base64 = base64.b64encode(archivo_binario.read()).decode("utf-8")
                    datos_payload = {"message": f"Sincronizacion MCE Planta ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})", "content": excel_base64, "branch": "main", "committer": {"name": usuario_git, "email": email_git}}
                    if sha_archivo: datos_payload["sha"] = sha_archivo
                    respuesta_put = requests.put(url_api, headers=cabeceras, data=json.dumps(datos_payload))
                    if respuesta_put.status_code in (200, 201):
                        st.success("✅ ¡Éxito Absoluto! Base respaldada directamente en GitHub."); st.balloons(); st.rerun()
                    else: st.error(f"❌ Error en la API. Respuesta: {respuesta_put.text}")
                except Exception as error_global_api: st.error(f"Fallo critico HTTP REST: {error_global_api}")
        st.write("---")
        t1, t2, t3 = st.tabs(["➕ Altas Catálogos", "✏️ Tabla de Edición Directa", "📥 Carga Masiva Excel"])
        with t1:
            n_n = st.text_input("Nombre de Colaborador:")
            if st.button("Registrar Empleado") and n_n: 
                st.session_state.personal[n_n] = None
                guardar_catalogos(st.session_state.personal, st.session_state.areas)
                st.success("Registrado."); st.rerun()
        with t2:
            st.subheader("✏️ Edición en Caliente de la Matriz MCE")
            st.warning("⚠️ Nota: Solo el Administrador está autorizado para borrar registros físicamente en la matriz.")
            df_editable = pd.DataFrame(st.session_state.actividades)
            if not df_editable.empty:
                configuracion_columnas = {
                    "No": st.column_config.NumberColumn("No", disabled=True, format="%d"),
                    "Origen": st.column_config.SelectboxColumn("Origen", options=LISTA_CLASIFICACIONES, required=True),
                    "Prioridad": st.column_config.SelectboxColumn("Prioridad", options=["Baja", "Media", "Urgente"], required=True),
                    "Responsable": st.column_config.SelectboxColumn("Responsable", options=list(st.session_state.personal.keys())),
                    "Area": st.column_config.SelectboxColumn("Área", options=st.session_state.areas),
                    "Fecha Inicio": st.column_config.TextColumn("Fecha Inicio"),
                    "Descripcion": st.column_config.TextColumn("Descripción", width="large"),
                    "% Avance": st.column_config.NumberColumn("% Avance", min_value=0, max_value=100, format="%d%%"),
                    "Fecha Compromiso": st.column_config.TextColumn("Fecha Compromiso"),
                    "Comentario": st.column_config.TextColumn("Comentario", width="medium"),
                    "Evidencia": st.column_config.TextColumn("Ruta Evidencia", disabled=True)
                }
                df_modificado = st.data_editor(df_editable, column_config=configuracion_columnas, use_container_width=True, hide_index=True, num_rows="dynamic", key="editor_tabla_master")
                if st.button("💾 CONFIRMAR Y GUARDAR CAMBIOS EN LA MATRIZ", type="primary", use_container_width=True):
                    st.session_state.actividades = df_modificado
                    try:
                        st.session_state.actividades.to_excel(ARCHIVO_DB, index=False)
                        st.success("✅ ¡Sincronizado!"); st.rerun()
                    except Exception as e_master: st.error(f"Error: {e_master}")
            else: st.info("No hay registros.")
        with t3:
            st.subheader("Inyección de Datos por Carga Masiva")
            columnas_p = ["Origen", "Fecha Inicio", "Prioridad", "Responsable", "Area", "Descripcion", "% Avance", "Fecha Compromiso", "Comentario"]
            ej_g = pd.DataFrame([["Programa de Actividades", datetime.now().strftime("%d-%b-%y"), "Media", "Bryan Flores", "⚙️ Ingeniería", "Descripción...", 0, "15-Jun-26", "..."]], columns=columnas_p)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as w:
                ej_g.to_excel(w, index=False, sheet_name='Plantilla')
                ws = w.sheets['Plantilla']
                anchos = {'A': 25, 'B': 15, 'C': 15, 'D': 22, 'E': 15, 'F': 45, 'G': 12, 'H': 20, 'I': 25}
                for col, ancho in anchos.items(): ws.column_dimensions[col].width = ancho
            st.download_button(label="📥 Descargar Plantilla Oficial (.xlsx)", data=buf.getvalue(), file_name="Plantilla_MCE.xlsx", key="btn_download_template_master")
            st.write("---")
            ex = st.file_uploader("Subir Excel modificado", type=["xlsx"], key="uploader_bulk_master")
            if ex is not None:
                df_ex = pd.read_excel(ex)
                if st.button("Confirmar Importación Masiva"):
                    if "No" not in df_ex.columns: df_ex.insert(0, "No", range(st.session_state.actividades["No"].max() + 1 if not st.session_state.actividades.empty else 1, (st.session_state.actividades["No"].max() + 1 if not st.session_state.actividades.empty else 1) + len(df_ex)))
                    if "Evidencia" not in df_ex.columns: df_ex["Evidencia"] = ""
                    st.session_state.actividades = pd.concat([st.session_state.actividades, df_ex], ignore_index=True)
                    try:
                        st.session_state.actividades.to_excel(ARCHIVO_DB, index=False)
                        st.success("¡Importado!"); st.rerun()
                    except Exception as e_b: st.error(f"Error: {e_b}")

    elif opcion_menu == "👑 Reglas de Liderazgo":
        st.markdown('<h2 style="color: #EC2024; font-weight: bold; margin-bottom: 20px;">👑 REGLAS DE LIDERAZGO: PLANTA METALES</h2>', unsafe_allow_html=True)
        st.write("Guía oficial de comportamiento, gestión y control en piso de producción.")
        st.write("---")

        col_l1, col_l2 = st.columns(2)

        with col_l1:
            with st.expander("🪵 Liderazgo y Comportamiento en Piso", expanded=True):
                st.markdown("""
                1. **Predica con el ejemplo:** Cumple las normas de seguridad, puntualidad y vestimenta antes que nadie.
                2. **Lidera desde el Gemba:** Pasa el 80% de tu tiempo en el piso de producción, no en la oficina.
                3. **Corrige en privado, reconoce en público:** Protege la dignidad de tu personal y celebra sus éxitos frente al grupo.
                4. **Escucha antes de juzgar:** Investiga la causa raíz de un problema antes de buscar culpables.
                5. **Mantén la consistencia:** Aplica las reglas y sanciones de manera justa, equitativa y sin favoritismos.
                """, unsafe_allow_html=True)

            with st.expander("💼 Operación y Gestión del Personal", expanded=True):
                st.markdown("""
                1. **Cero tolerancia al ocio:** Asegura que cada operador tenga una tarea asignada en todo momento.
                2. **Activa el plan B de inmediato:** Si la producción se detiene, reasigna al personal a actividades secundarias (5S, capacitación, mantenimiento).
                3. **Informa desviaciones a Dirección:** Reporta al final del turno cualquier falta de actividad programada y justifica el uso del tiempo.
                4. **Capacita constantemente:** Desarrollar la polivalencia de tu equipo reduce la dependencia de personas específicas.
                5. **La seguridad es primero:** Detén cualquier operación que ponga en riesgo la integridad física del personal.
                """, unsafe_allow_html=True)

        with col_l2:
            with st.expander("📢 Comunicación y Relaciones Interdepartamentales", expanded=True):
                st.markdown("""
                1. **Arranca con juntas Tier 1:** Realiza reuniones de 5 minutos al inicio del turno para alinear metas y riesgos.
                2. **Comunícate con datos:** Al solicitar apoyo a Calidad o Mantenimiento, habla con números y hechos, no con opiniones.
                3. **Define el 'para cuándo':** Al delegar o solicitar una tarea a otra jefatura, establece siempre una fecha y hora límite.
                4. **Cierra el ciclo de comunicación:** Confirma que el receptor entendió el mensaje pidiéndole que lo explique con sus palabras.
                5. **Fomenta las ideas de mejora:** Escucha y canaliza las propuestas de los operadores para optimizar los procesos.
                """, unsafe_allow_html=True)

            with st.expander("🎯 Delegación y Control", expanded=True):
                st.markdown("""
                1. **Delega con base en la habilidad:** Asigna las tareas críticas al personal que ya demostró la competencia para resolverlas.
                2. **Sigue el rastro, no microgestiones:** Define puntos de revisión intermedios en lugar de vigilar cada paso del proceso.
                3. **Entrega recursos completos:** Al delegar, asegura que la persona tenga la herramienta, la información y el tiempo necesarios.
                4. **Asume la responsabilidad final:** Si tu equipo falla, tú eres el responsable ante la Dirección; no culpes a tus subordinados.
                5. **Estandariza los éxitos:** Cuando una solución funcione, documenta el nuevo método para que se convierta en la regla oficial.
                """, unsafe_allow_html=True)

    elif opcion_menu == "📋 Reportes PDF":
        st.subheader("🛠️ Generación de Reportes Ejecutivos")
        st.write("Selecciona los filtros requeridos para estructurar los reportes de la planta:")
        st.write("---")
        
        col_rep1, col_rep2, col_rep3 = st.columns(3)
        with col_rep1: 
            area_rep = st.selectbox("Filtrar por Área", ["Todas"] + st.session_state.areas, key="pdf_area")
        with col_rep2: 
            resp_rep = st.selectbox("Filtrar por Responsable", ["Todos"] + list(st.session_state.personal.keys()), key="pdf_resp")
        with col_rep3:
            rango_tiempo = st.selectbox(
                "Rango de Fecha Compromiso", 
                ["Cualquier Fecha Límite", "Esta Semana", "Próximas 2 Semanas", "Próximas 3 Semanas", "Próximas 4 Semanas"], 
                key="pdf_tiempo"
            )
        df_rep = pd.DataFrame(st.session_state.actividades)
        if not df_rep.empty:
            if area_rep != "Todas": 
                df_rep = df_rep[df_rep["Area"] == area_rep]
            if resp_rep != "Todos": 
                df_rep = df_rep[df_rep["Responsable"] == resp_rep]
                
            df_todas_pendientes = df_rep[df_rep["% Avance"].astype(int) < 100].copy()
            df_todas_terminadas = df_rep[df_rep["% Avance"].astype(int) == 100].copy()
            
            if rango_tiempo != "Cualquier Fecha Límite":
                hoy = datetime.now().date()
                if rango_tiempo == "Esta Semana": dias_limite = 7
                elif rango_tiempo == "Próximas 2 Semanas": dias_limite = 14
                elif rango_tiempo == "Próximas 3 Semanas": dias_limite = 21
                elif rango_tiempo == "Próximas 4 Semanas": dias_limite = 28
                    
                fecha_maxima = hoy + timedelta(days=dias_limite)
                
                def evaluar_rango_fecha(fecha_str):
                    try:
                        if not fecha_str or str(fecha_str).strip() in ["None", "nan", ""]: return False
                        f_dt = datetime.strptime(str(fecha_str).strip(), "%d-%b-%y").date()
                        return f_dt <= fecha_maxima
                    except: return False
                        
                if not df_todas_pendientes.empty:
                    df_todas_pendientes = df_todas_pendientes[df_todas_pendientes["Fecha Compromiso"].apply(evaluar_rango_fecha)]
                if not df_todas_terminadas.empty:
                    df_todas_terminadas = df_todas_terminadas[df_todas_terminadas["Fecha Compromiso"].apply(evaluar_rango_fecha)]
                    
            st.write("### 📊 Vista Previa en Plataforma")
            c_p1, c_p2 = st.columns(2)
            c_p1.metric("Tareas Pendientes Encontradas", len(df_todas_pendientes))
            c_p2.metric("Tareas Terminadas Encontradas", len(df_todas_terminadas))
            
            if not df_todas_pendientes.empty or not df_todas_terminadas.empty:
                if not df_todas_pendientes.empty:
                    st.write("**Pendientes Actuales:**")
                    st.dataframe(df_todas_pendientes[["No", "Responsable", "Area", "Descripcion", "% Avance", "Fecha Compromiso"]], use_container_width=True, hide_index=True)
                if not df_todas_terminadas.empty:
                    st.write("**Terminadas Recientes:**")
                    st.dataframe(df_todas_terminadas[["No", "Responsable", "Area", "Descripcion", "% Avance", "Fecha Compromiso"]], use_container_width=True, hide_index=True)
                st.write("---")
                
                def creador_documento_pdf(df_datos, es_bloque_pendientes, area_txt, resp_txt, tiempo_txt):
                    from fpdf import FPDF
                    pdf = FPDF(orientation="L", unit="mm", format="A4")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.add_page()
                    
                    pdf.set_font("Helvetica", "B", 16)
                    pdf.set_text_color(12, 35, 64)
                    pdf.cell(0, 10, txt="PLANTA METALES Y MAQUINADOS", ln=True, align="C")
                    pdf.set_font("Helvetica", "B", 12)
                    
                    tipo_reporte = "REPORTE DE ACTIVIDADES PENDIENTES" if es_bloque_pendientes else "REPORTE DE ACTIVIDADES TERMINADAS"
                    pdf.cell(0, 8, txt=tipo_reporte, ln=True, align="C")
                    
                    area_txt_limpio = limpiar_para_pdf(area_txt)
                    resp_txt_limpio = limpiar_para_pdf(resp_txt)
                    tiempo_txt_limpio = limpiar_para_pdf(tiempo_txt)
                    
                    pdf.set_font("Helvetica", "I", 9)
                    pdf.set_text_color(80, 80, 80)
                    pdf.cell(0, 5, txt=f"Filtros: Area ({area_txt_limpio}) | Responsable ({resp_txt_limpio}) | Periodo ({tiempo_txt_limpio})", ln=True, align="C")
                    pdf.cell(0, 5, txt=f"Generado el: {datetime.now().strftime('%d-%b-%y %H:%M')}", ln=True, align="C")
                    pdf.ln(3)
                    
                    pdf.set_draw_color(12, 35, 64)
                    pdf.line(10, pdf.get_y(), 285, pdf.get_y())
                    pdf.ln(4)
                    
                    groups = df_datos.groupby("Area")
                    for area_name, gp in groups:
                        a_limpio = limpiar_para_pdf(area_name)
                        pdf.set_font("Helvetica", "B", 10.5)
                        pdf.set_text_color(12, 35, 64)
                        pdf.cell(0, 6, txt=f">> AREA: {a_limpio.upper()}", ln=True)
                        
                        pdf.set_font("Helvetica", "B", 9)
                        pdf.set_fill_color(207, 216, 220)
                        pdf.cell(15, 6, txt="No", border=1, fill=True, align="C")
                        pdf.cell(50, 6, txt="Responsable", border=1, fill=True)
                        pdf.cell(145, 6, txt="Descripción de la Actividad", border=1, fill=True)
                        
                        col_estatus = "Avance" if es_bloque_pendientes else "Estatus"
                        pdf.cell(25, 6, txt=col_estatus, border=1, fill=True, align="C")
                        
                        col_fecha = "Fecha Límite" if es_bloque_pendientes else "Fecha Cierre"
                        pdf.cell(40, 6, txt=col_fecha, border=1, fill=True, ln=True)
                        
                        pdf.set_font("Helvetica", "", 8.5)
                        pdf.set_text_color(0, 0, 0)
                        for _, fila in gp.iterrows():
                            desc_cruda = str(fila["Descripcion"])
                            d_corta = desc_cruda[:82] + "..." if len(desc_cruda) > 85 else desc_cruda
                            r_limpio = limpiar_para_pdf(fila["Responsable"])
                            d_corta = limpiar_para_pdf(d_corta)
                            
                            if es_bloque_pendientes:
                                pdf.set_fill_color(255, 243, 205)
                                txt_avance = f"{fila['% Avance']}%"
                            else:
                                pdf.set_fill_color(212, 237, 218)
                                txt_avance = "Listo 100%"
                            
                            pdf.cell(15, 6, txt=str(fila["No"]), border=1, align="C")
                            pdf.cell(50, 6, txt=r_limpio[:25], border=1)
                            pdf.cell(145, 6, txt=d_corta, border=1)
                            pdf.cell(25, 6, txt=txt_avance, border=1, fill=True, align="C")
                            pdf.cell(40, 6, txt=str(fila["Fecha Compromiso"]), border=1, ln=True)
                        pdf.ln(4)
                        
                    return pdf.output()

                st.write("### 📥 Descarga de Archivos")
                col_btn1, col_btn2 = st.columns(2)
                
                with col_btn1:
                    if not df_todas_pendientes.empty:
                        try:
                            bytes_pend = creador_documento_pdf(df_todas_pendientes, True, area_rep, resp_rep, rango_tiempo)
                            st.download_button(
                                label="⚠️ Descargar Reporte de Pendientes (PDF)",
                                data=bytes(bytes_pend),
                                file_name=f"Reporte_PENDIENTES_{datetime.now().strftime('%d-%b-%y')}.pdf",
                                mime="application/pdf",
                                use_container_width=True
                            )
                        except Exception as e_p:
                            st.error(f"Error en PDF de pendientes: {e_p}")
                    else:
                        st.info("No hay tareas pendientes para exportar con estos filtros.")
                        
                with col_btn2:
                    if not df_todas_terminadas.empty:
                        try:
                            bytes_term = creador_documento_pdf(df_todas_terminadas, False, area_rep, resp_rep, rango_tiempo)
                            st.download_button(
                                label="✅ Descargar Reporte de Terminadas (PDF)",
                                data=bytes(bytes_term),
                                file_name=f"Reporte_TERMINADAS_{datetime.now().strftime('%d-%b-%y')}.pdf",
                                mime="application/pdf",
                                use_container_width=True
                            )
                        except Exception as e_t:
                            st.error(f"Error en PDF de terminadas: {e_t}")
                    else:
                        st.info("No hay tareas terminadas para exportar con estos filtros.")
            else:
                st.success("🎉 ¡Sin movimientos! No hay actividades registradas en los parámetros seleccionados.")
        else:
            st.info("La base de datos se encuentra vacía.")

    elif opcion_menu == "🏭 Industria 4.0 & Stack":
        st.markdown('<h2 style="color: #EC2024; font-weight: bold; margin-bottom: 5px; font-family: \'Montserrat\', sans-serif;">🏭 MANUFACTURA INTELIGENTE E INDUSTRIA 4.0</h2>', unsafe_allow_html=True)
        st.markdown('<p style="font-family: \'Montserrat\', sans-serif; font-size: 16px; font-weight: bold; color: #111111; margin-bottom: 20px; text-transform: uppercase;">SOLUCIONES QUE TRANSFORMAN TU EMPRESA</p>', unsafe_allow_html=True)
        st.write("---")
        
        col_ind1, col_ind2 = st.columns(2)
        
        with col_ind1:
            with st.expander("📝 Justificación de Manufactura Inteligente", expanded=True):
                st.markdown("""
                En la era de la **Industria 4.0**, la digitalización de piso de producción (**Gemba**) es clave para maximizar la productividad y responder ágilmente a contingencias. 
                La **Matriz de Comunicación Efectiva (MCE)** de Planta Metales y Maquinados sustituye los flujos de papel y hojas de cálculo locales por una plataforma web reactiva y centralizada.
                
                *   **Eliminación de Silos de Información**: Toda la planta visualiza el mismo estado en tiempo real, desde operadores hasta la dirección.
                *   **Reducción de Tiempos Muertos**: La comunicación instantánea de desviaciones agiliza la respuesta de mantenimiento, calidad e ingeniería.
                *   **Decisiones Basadas en Datos**: Reemplaza opiniones subjetivas con gráficos de Pareto de tareas y tiempos de vencimiento reales.
                """)
                
            with st.expander("🚀 Beneficios Estratégicos del Proyecto", expanded=True):
                st.markdown("""
                *   **Trazabilidad y Calidad**: El cierre de tareas (100%) requiere evidencia fotográfica obligatoria, asegurando la verificación física de los trabajos en piso.
                *   **Sincronización en la Nube**: Integración automática mediante la API de GitHub para respaldar el archivo maestro Excel de forma segura e inmediata.
                *   **Monitoreo de Líderes**: Identificación visual del avance promedio y eficiencia por responsable directo para equilibrar la carga de trabajo.
                *   **Cero Papel**: Reducción de costos de impresión y archivado físico, apoyando iniciativas sustentables.
                """)
                
        with col_ind2:
            with st.expander("🛠️ Resumen del Stack Tecnológico", expanded=True):
                st.markdown("""
                La aplicación está construida sobre un stack tecnológico moderno, ligero y de alta velocidad:
                
                1.  **Frontend & Logic**: **Streamlit (Python)**, que permite interfaces dinámicas y renderización reactiva de datos.
                2.  **Base de Datos**: **Pandas + Excel (Openpyxl)**, facilitando la edición interactiva directa en piso y la persistencia de datos tradicional.
                3.  **Visualizaciones**: **Plotly**, generando gráficos de Pareto e indicadores de rendimiento interactivos.
                4.  **Generación de PDF**: **FPDF2**, optimizado para reportes corporativos estructurados con soporte de caracteres en español.
                5.  **Control de Versiones & API**: **GitHub API REST**, respaldando la matriz Excel en la nube de forma transparente y encriptada.
                """)
                
            with st.expander("🔐 Guía de Acceso (Usuarios y Contraseñas)", expanded=True):
                st.markdown("""
                A continuación se detallan los perfiles autorizados para operar la plataforma:
                
                *   **Colaboradores / Operadores (Lectura y Avances)**:
                    *   *Usuarios*: Jesús Morales, Cruz Carreón, Luis Quintana, Bryan Flores, Rodolfo Fernández M., etc.
                    *   *Acceso*: Libre. Permite ajustar porcentaje de avances en barra de progreso rápida, comentar la bitácora y subir fotos de evidencia.
                *   **Carga de Actividades (Usuario)**:
                    *   *Acceso*: Protegido por contraseña.
                    *   *Contraseña*: `Metales`
                    *   *Permisos*: Crear y registrar nuevas actividades en el catálogo.
                *   **Administrador (Máster)**:
                    *   *Acceso*: Protegido por contraseña máster.
                    *   *Contraseña*: `SigramaMetales2026`
                    *   *Permisos*: Registro de nuevos colaboradores, importación de plantilla masiva, sincronización manual/respaldo en GitHub y **edición en caliente / borrado físico de registros** (Solo el Administrador cuenta con estos permisos).
                """)
